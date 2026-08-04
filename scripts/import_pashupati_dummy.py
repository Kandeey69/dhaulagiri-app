from __future__ import annotations

import json
import os
import re
import shutil
import sqlite3
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
BACKUP_PATH = Path(
    r"C:\Users\Kandeey\Desktop\export data for import\Pashupati-Chemicals-2082-83-backup.easysolution-backup.json"
)
WORKBOOK_PATH = Path(r"C:\Users\Kandeey\Documents\Pashupati stock.xlsx")
APPDATA_DIR = Path(os.environ["APPDATA"]) / "com.easysolution.businesssuite"
BACKUP_DIR = ROOT / "data-load-backups" / datetime.now().strftime("%Y%m%d-%H%M%S")


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace(",", "").strip()
    return float(cleaned) if cleaned else 0.0


def clean_bill_no(value: Any) -> str:
    return text(value)


def clean_item_name(value: Any) -> str:
    item = re.sub(r"\s+", " ", text(value)).strip()
    return item.replace("( ", "(").replace(" )", ")")


def clean_date(value: Any) -> str:
    raw = text(value).replace("-", "/")
    match = re.match(r"^(\d{4})/(\d{1,2})/(\d{1,2})$", raw)
    if not match:
        return raw
    year, month, day = match.groups()
    return f"{year}/{int(month):02d}/{int(day):02d}"


def fiscal_year_id(company_id: str, code: str) -> str:
    company_part = re.sub(r"[^a-zA-Z0-9_-]+", "-", company_id or "default")
    code_part = re.sub(r"[^a-zA-Z0-9]+", "-", code or "legacy")
    return f"{company_part}-{code_part}"


def stock_database_name(company_id: str) -> str:
    if not company_id or company_id == "default":
        return "inventorytracked-stock.db"
    encoded = "-".join(f"{ord(ch):04x}" for ch in company_id.strip())
    return f"inventorytracked-stock-{encoded}.db"


def deterministic_id(kind: str, *parts: Any) -> str:
    raw = "|".join([kind, *(text(part) for part in parts)])
    return str(uuid.uuid5(uuid.NAMESPACE_URL, raw))


def connect(path: Path) -> sqlite3.Connection:
    path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA busy_timeout = 10000")
    conn.execute("PRAGMA journal_mode = WAL")
    conn.execute("PRAGMA synchronous = NORMAL")
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def backup_existing_appdata() -> None:
    APPDATA_DIR.mkdir(parents=True, exist_ok=True)
    files = [
        "accounts-pashupati-chemicals-2082-83.db",
        "import-purchases-pashupati-chemicals-2082-83.db",
        stock_database_name("pashupati-chemicals-2082-83"),
        "company-profiles.seed.json",
    ]
    if any((APPDATA_DIR / file_name).exists() for file_name in files):
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)
        for file_name in files:
            for suffix in ("", "-wal", "-shm"):
                source = APPDATA_DIR / f"{file_name}{suffix}"
                if source.exists():
                    shutil.copy2(source, BACKUP_DIR / source.name)
                    source.unlink()


def create_accounts_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS fiscal_years (
          id TEXT PRIMARY KEY,
          companyId TEXT NOT NULL,
          code TEXT NOT NULL,
          startBs TEXT NOT NULL,
          endBs TEXT NOT NULL,
          startAd TEXT NOT NULL DEFAULT '',
          endAd TEXT NOT NULL DEFAULT '',
          status TEXT NOT NULL DEFAULT 'OPEN',
          createdAt TEXT NOT NULL,
          updatedAt TEXT NOT NULL,
          UNIQUE(companyId, code)
        );
        CREATE TABLE IF NOT EXISTS receipt_allocations (
          id TEXT PRIMARY KEY,
          receipt_id TEXT NOT NULL,
          sale_id TEXT NOT NULL,
          amount_npr REAL NOT NULL DEFAULT 0,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS ledger_entries (
          id TEXT PRIMARY KEY,
          batch_id TEXT NOT NULL,
          company_id TEXT NOT NULL,
          fiscal_year_id TEXT NOT NULL,
          transaction_date TEXT NOT NULL,
          account_code TEXT NOT NULL,
          party_id TEXT,
          source_type TEXT NOT NULL,
          source_id TEXT NOT NULL,
          posting_version TEXT NOT NULL DEFAULT 'v1',
          debit REAL NOT NULL DEFAULT 0,
          credit REAL NOT NULL DEFAULT 0,
          narration TEXT NOT NULL DEFAULT '',
          status TEXT NOT NULL DEFAULT 'ACTIVE',
          reversal_of_entry_id TEXT,
          created_at TEXT NOT NULL,
          updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS parties (
          id TEXT PRIMARY KEY,
          name TEXT NOT NULL UNIQUE,
          address TEXT,
          phone TEXT,
          pan_no TEXT,
          opening_balance REAL DEFAULT 0,
          is_active INTEGER DEFAULT 1,
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sales (
          id TEXT PRIMARY KEY,
          bill_no TEXT NOT NULL UNIQUE,
          date_bs TEXT,
          date_ad TEXT,
          party_id TEXT NOT NULL,
          quantity REAL DEFAULT 0,
          rate REAL DEFAULT 0,
          amount REAL NOT NULL,
          sales_amount REAL DEFAULT 0,
          vat_amount REAL DEFAULT 0,
          total_amount REAL DEFAULT 0,
          remarks TEXT,
          created_at TEXT NOT NULL,
          fiscal_year_id TEXT NOT NULL DEFAULT '',
          lifecycle_status TEXT NOT NULL DEFAULT 'POSTED',
          applied_vat_rate REAL NOT NULL DEFAULT 13,
          calculation_version TEXT NOT NULL DEFAULT 'sales-policy-v1',
          calculated_at TEXT NOT NULL DEFAULT '',
          posted_at TEXT NOT NULL DEFAULT '',
          posted_by TEXT NOT NULL DEFAULT '',
          voided_at TEXT NOT NULL DEFAULT '',
          reversed_at TEXT NOT NULL DEFAULT '',
          reversal_reason TEXT NOT NULL DEFAULT '',
          replacement_transaction_id TEXT NOT NULL DEFAULT ''
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_sales_bill_no_whole ON sales(bill_no);
        CREATE TABLE IF NOT EXISTS collections (
          id TEXT PRIMARY KEY,
          date_bs TEXT,
          date_ad TEXT,
          party_id TEXT NOT NULL,
          bank_name TEXT,
          amount REAL NOT NULL,
          reference_no TEXT,
          remarks TEXT,
          created_at TEXT NOT NULL,
          fiscal_year_id TEXT NOT NULL DEFAULT '',
          lifecycle_status TEXT NOT NULL DEFAULT 'POSTED',
          posted_at TEXT NOT NULL DEFAULT '',
          posted_by TEXT NOT NULL DEFAULT '',
          voided_at TEXT NOT NULL DEFAULT '',
          reversed_at TEXT NOT NULL DEFAULT '',
          reversal_reason TEXT NOT NULL DEFAULT '',
          replacement_transaction_id TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS credit_notes (
          id TEXT PRIMARY KEY,
          credit_note_no TEXT NOT NULL UNIQUE,
          date_bs TEXT,
          date_ad TEXT,
          party_id TEXT NOT NULL,
          amount REAL NOT NULL,
          vat_amount REAL DEFAULT 0,
          total_amount REAL DEFAULT 0,
          remarks TEXT,
          created_at TEXT NOT NULL,
          fiscal_year_id TEXT NOT NULL DEFAULT '',
          lifecycle_status TEXT NOT NULL DEFAULT 'POSTED',
          posted_at TEXT NOT NULL DEFAULT '',
          posted_by TEXT NOT NULL DEFAULT '',
          voided_at TEXT NOT NULL DEFAULT '',
          reversed_at TEXT NOT NULL DEFAULT '',
          reversal_reason TEXT NOT NULL DEFAULT '',
          replacement_transaction_id TEXT NOT NULL DEFAULT ''
        );
        CREATE TABLE IF NOT EXISTS activity_logs (
          id TEXT PRIMARY KEY,
          action TEXT NOT NULL,
          detail TEXT NOT NULL,
          created_at TEXT NOT NULL,
          metadata TEXT NOT NULL DEFAULT ''
        );
        CREATE INDEX IF NOT EXISTS idx_sales_party ON sales(party_id);
        CREATE INDEX IF NOT EXISTS idx_sales_fiscal_year ON sales(fiscal_year_id);
        CREATE INDEX IF NOT EXISTS idx_collections_party ON collections(party_id);
        CREATE INDEX IF NOT EXISTS idx_collections_fiscal_year ON collections(fiscal_year_id);
        CREATE INDEX IF NOT EXISTS idx_credit_notes_party ON credit_notes(party_id);
        CREATE INDEX IF NOT EXISTS idx_credit_notes_fiscal_year ON credit_notes(fiscal_year_id);
        CREATE INDEX IF NOT EXISTS idx_receipt_allocations_receipt ON receipt_allocations(receipt_id);
        CREATE INDEX IF NOT EXISTS idx_receipt_allocations_sale ON receipt_allocations(sale_id);
        CREATE INDEX IF NOT EXISTS idx_ledger_source ON ledger_entries(source_type, source_id, posting_version);
        CREATE INDEX IF NOT EXISTS idx_ledger_fiscal_year ON ledger_entries(fiscal_year_id);
        CREATE INDEX IF NOT EXISTS idx_ledger_party ON ledger_entries(party_id);
        """
    )


def insert_fiscal_year(conn: sqlite3.Connection, company_id: str, code: str) -> str:
    fy_id = fiscal_year_id(company_id, code)
    timestamp = now_iso()
    start_year = int(code.split("/")[0]) if "/" in code else 2082
    conn.execute(
        """
        INSERT OR REPLACE INTO fiscal_years
          (id, companyId, code, startBs, endBs, startAd, endAd, status, createdAt, updatedAt)
        VALUES (?, ?, ?, ?, ?, '', '', 'OPEN', ?, ?)
        """,
        (fy_id, company_id, code, f"{start_year}/04/01", f"{start_year + 1}/03/32", timestamp, timestamp),
    )
    return fy_id


def load_accounts(conn: sqlite3.Connection, accounts: dict[str, Any], company_id: str, fy_code: str) -> None:
    create_accounts_schema(conn)
    fy_id = insert_fiscal_year(conn, company_id, fy_code)
    for party in accounts.get("parties", []):
        conn.execute(
            """
            INSERT INTO parties (id, name, address, phone, pan_no, opening_balance, is_active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                party["id"],
                party["name"],
                party.get("address", ""),
                party.get("phone", ""),
                party.get("panNo", ""),
                number(party.get("openingBalance", 0)),
                1 if party.get("isActive", True) else 0,
                party.get("createdAt", now_iso()),
            ),
        )
    for sale in accounts.get("sales", []):
        conn.execute(
            """
            INSERT INTO sales (
              id, bill_no, date_bs, date_ad, party_id, quantity, rate, amount, sales_amount,
              vat_amount, total_amount, remarks, created_at, fiscal_year_id, lifecycle_status,
              applied_vat_rate, calculation_version, calculated_at, posted_at, posted_by,
              voided_at, reversed_at, reversal_reason, replacement_transaction_id
            )
            VALUES (?, ?, ?, ?, ?, 0, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                sale["id"],
                sale["billNo"],
                sale.get("dateBs", ""),
                sale.get("dateAd", ""),
                sale["partyId"],
                number(sale.get("totalAmount", sale.get("amount", 0))),
                number(sale.get("salesAmount", 0)),
                number(sale.get("vatAmount", 0)),
                number(sale.get("totalAmount", sale.get("amount", 0))),
                sale.get("remarks", ""),
                sale.get("createdAt", now_iso()),
                sale.get("fiscalYearId") or fy_id,
                sale.get("lifecycleStatus", "POSTED"),
                number(sale.get("appliedVatRate", 13)),
                sale.get("calculationVersion", "sales-policy-v1"),
                sale.get("calculatedAt", ""),
                sale.get("postedAt", ""),
                sale.get("postedBy", ""),
                sale.get("voidedAt", ""),
                sale.get("reversedAt", ""),
                sale.get("reversalReason", ""),
                sale.get("replacementTransactionId", ""),
            ),
        )
    for collection in accounts.get("collections", []):
        conn.execute(
            """
            INSERT INTO collections (
              id, date_bs, date_ad, party_id, bank_name, amount, reference_no, remarks, created_at,
              fiscal_year_id, lifecycle_status, posted_at, posted_by, voided_at, reversed_at,
              reversal_reason, replacement_transaction_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                collection["id"],
                collection.get("dateBs", ""),
                collection.get("dateAd", ""),
                collection["partyId"],
                collection.get("bankName", ""),
                number(collection.get("amount", 0)),
                collection.get("referenceNo") or collection.get("receiptNo", ""),
                collection.get("remarks", ""),
                collection.get("createdAt", now_iso()),
                collection.get("fiscalYearId") or fy_id,
                collection.get("lifecycleStatus", "POSTED"),
                collection.get("postedAt", ""),
                collection.get("postedBy", ""),
                collection.get("voidedAt", ""),
                collection.get("reversedAt", ""),
                collection.get("reversalReason", ""),
                collection.get("replacementTransactionId", ""),
            ),
        )
    for note in accounts.get("creditNotes", []):
        conn.execute(
            """
            INSERT INTO credit_notes (
              id, credit_note_no, date_bs, date_ad, party_id, amount, vat_amount, total_amount,
              remarks, created_at, fiscal_year_id, lifecycle_status, posted_at, posted_by,
              voided_at, reversed_at, reversal_reason, replacement_transaction_id
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                note["id"],
                note.get("creditNoteNo", note["id"]),
                note.get("dateBs", ""),
                note.get("dateAd", ""),
                note["partyId"],
                number(note.get("amount", 0)),
                number(note.get("vatAmount", 0)),
                number(note.get("totalAmount", note.get("amount", 0))),
                note.get("remarks", ""),
                note.get("createdAt", now_iso()),
                note.get("fiscalYearId") or fy_id,
                note.get("lifecycleStatus", "POSTED"),
                note.get("postedAt", ""),
                note.get("postedBy", ""),
                note.get("voidedAt", ""),
                note.get("reversedAt", ""),
                note.get("reversalReason", ""),
                note.get("replacementTransactionId", ""),
            ),
        )
    for log in accounts.get("activityLogs", []):
        conn.execute(
            "INSERT INTO activity_logs (id, action, detail, created_at, metadata) VALUES (?, ?, ?, ?, ?)",
            (
                log["id"],
                log.get("action", ""),
                log.get("detail", log.get("details", "")),
                log.get("createdAt", now_iso()),
                json.dumps(log.get("metadata", "")),
            ),
        )
    conn.commit()


def load_purchase(conn: sqlite3.Connection, purchase: dict[str, Any], company_id: str, fy_code: str) -> None:
    conn.executescript((ROOT / "src" / "purchase" / "db" / "schema.sql").read_text(encoding="utf-8"))
    fy_id = insert_fiscal_year(conn, company_id, fy_code)
    settings = purchase.get("settings") or {}
    conn.execute(
        """
        INSERT OR REPLACE INTO app_settings (
          id, companyName, fiscalYear, defaultExchangeRate, supplierPurchaseCurrency,
          panVatNo, address, phone, agentServiceVatRate
        )
        VALUES ('default', ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            settings.get("companyName", "Pashupati Chemicals"),
            settings.get("fiscalYear", fy_code),
            number(settings.get("defaultExchangeRate", 1.6015)),
            settings.get("supplierPurchaseCurrency", "INR"),
            settings.get("panVatNo", ""),
            settings.get("address", ""),
            settings.get("phone", ""),
            number(settings.get("agentServiceVatRate", 13)),
        ),
    )
    for party in purchase.get("parties", []):
        conn.execute(
            """
            INSERT INTO parties (
              id, name, address, phone, panVatNo, country, category, openingPayable,
              isActive, createdAt, updatedAt
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                party["id"],
                party["name"],
                party.get("address", ""),
                party.get("phone", ""),
                party.get("panVatNo", ""),
                party.get("country", ""),
                party.get("category", "Supplier"),
                number(party.get("openingPayable", 0)),
                1 if party.get("isActive", True) else 0,
                party.get("createdAt", now_iso()),
                party.get("updatedAt", party.get("createdAt", now_iso())),
            ),
        )
    purchase_cols = [
        "id", "fiscalYearId", "lifecycleStatus", "vendorPartyId", "vendorBillNumber", "billDate",
        "supplierCurrency", "amountIC", "supplierExchangeRate", "supplierAmountNPR", "customAgentPartyId",
        "debitNoteNumber", "debitNoteDate", "importDutyNPR", "customServiceNPR", "importVatNPR",
        "terminalChargeWithoutVatNPR", "terminalVatNPR", "totalTerminalChargeNPR", "freightIndiaStatus",
        "freightIndiaPartyId", "freightIndiaAmountIC", "freightIndiaExchangeRate", "freightIndiaAmountNPR",
        "totalKg", "loadingUnloadingChargePerKg", "loadingUnloadingChargeNPR", "otherChargesNPR",
        "debitNoteTotalNPR", "agentServiceBillNumber", "agentServiceBillDate", "agentServiceAmountBeforeVatNPR",
        "agentServiceVatNPR", "agentServiceTotalNPR", "totalAgentPayableNPR", "totalInputVatNPR",
        "landedCostNPR", "appliedVatRate", "appliedExchangeRate", "calculationVersion", "calculatedAt",
        "postedAt", "postedBy", "voidedAt", "reversedAt", "reversalReason", "replacementTransactionId",
        "remarks", "createdAt", "updatedAt",
    ]
    numeric_purchase_cols = {
        "amountIC", "supplierExchangeRate", "supplierAmountNPR", "importDutyNPR", "customServiceNPR",
        "importVatNPR", "terminalChargeWithoutVatNPR", "terminalVatNPR", "totalTerminalChargeNPR",
        "freightIndiaAmountIC", "freightIndiaExchangeRate", "freightIndiaAmountNPR", "totalKg",
        "loadingUnloadingChargePerKg", "loadingUnloadingChargeNPR", "otherChargesNPR", "debitNoteTotalNPR",
        "agentServiceAmountBeforeVatNPR", "agentServiceVatNPR", "agentServiceTotalNPR", "totalAgentPayableNPR",
        "totalInputVatNPR", "landedCostNPR", "appliedVatRate", "appliedExchangeRate",
    }
    purchase_defaults = {
        "fiscalYearId": fy_id,
        "lifecycleStatus": "POSTED",
        "supplierCurrency": "INR",
        "freightIndiaStatus": "Paid by custom agent",
        "appliedVatRate": 13,
        "appliedExchangeRate": 1.6015,
        "calculationVersion": "legacy-migrated-v1",
    }
    placeholders = ", ".join("?" for _ in purchase_cols)
    for row in purchase.get("purchases", []):
        values = []
        for col in purchase_cols:
            value = row.get(col, purchase_defaults.get(col, ""))
            if col in numeric_purchase_cols:
                value = number(value)
            values.append(value)
        conn.execute(
            f"INSERT INTO import_purchases ({', '.join(purchase_cols)}) VALUES ({placeholders})",
            values,
        )
    payment_cols = [
        "id", "fiscalYearId", "lifecycleStatus", "partyId", "paymentDate", "paymentType", "currency",
        "amount", "exchangeRate", "amountNPR", "paymentMethod", "referenceNumber", "remarks", "postedAt",
        "postedBy", "voidedAt", "reversedAt", "reversalReason", "replacementTransactionId", "createdAt", "updatedAt",
    ]
    for row in purchase.get("payments", []):
        values = []
        for col in payment_cols:
            value = row.get(col, "")
            if col == "fiscalYearId" and not value:
                value = fy_id
            if col == "lifecycleStatus" and not value:
                value = "POSTED"
            if col in {"amount", "exchangeRate", "amountNPR"}:
                value = number(value)
            values.append(value)
        conn.execute(f"INSERT INTO payments ({', '.join(payment_cols)}) VALUES ({', '.join('?' for _ in payment_cols)})", values)
    for log in purchase.get("activityLogs", []):
        conn.execute(
            """
            INSERT INTO activity_logs (id, action, details, userName, oldValue, newValue, metadata, createdAt)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                log["id"],
                log.get("action", ""),
                log.get("details", log.get("detail", "")),
                log.get("userName", ""),
                log.get("oldValue", ""),
                log.get("newValue", ""),
                json.dumps(log.get("metadata", "")),
                log.get("createdAt", now_iso()),
            ),
        )
    conn.commit()


def parse_workbook() -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    workbook = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    sales_lines: list[dict[str, Any]] = []
    purchase_lines: list[dict[str, Any]] = []
    ordered_items: list[str] = []
    seen_items: set[str] = set()

    for (item,) in workbook["stock type"].iter_rows(min_row=1, max_col=1, values_only=True):
        name = clean_item_name(item)
        if not name or name.lower().startswith(("item", "total", "grand", "note")):
            continue
        if "closing" in name.lower() or "purchase" in name.lower():
            continue
        if name not in seen_items:
            ordered_items.append(name)
            seen_items.add(name)

    for row in workbook["stock register"].iter_rows(min_row=2, values_only=True):
        date_bs, bill_no, party_name, item_name, qty, rate, total = row[:7]
        item = clean_item_name(item_name)
        bill = clean_bill_no(bill_no)
        if not bill or not item:
            continue
        sales_lines.append(
            {
                "date": clean_date(date_bs),
                "billNo": bill,
                "partyName": text(party_name),
                "itemName": item,
                "quantity": number(qty),
                "rate": number(rate),
                "amount": number(total),
            }
        )
        if item not in seen_items:
            ordered_items.append(item)
            seen_items.add(item)

    for row in workbook["Purchase register"].iter_rows(min_row=3, values_only=True):
        date_bs, bill_no, party_name, item_name, qty, inr_rate, inr_total, landed_npr = row[:8]
        item = clean_item_name(item_name)
        bill = clean_bill_no(bill_no)
        if not bill or not item:
            continue
        purchase_lines.append(
            {
                "date": clean_date(date_bs),
                "billNo": bill,
                "partyName": text(party_name),
                "itemName": item,
                "quantity": number(qty),
                "entryRate": number(inr_rate),
                "entryAmount": number(inr_total),
                "amount": number(landed_npr),
            }
        )
        if item not in seen_items:
            ordered_items.append(item)
            seen_items.add(item)

    return sales_lines, purchase_lines, ordered_items


def create_stock_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS stock_items (
          id TEXT PRIMARY KEY,
          code TEXT NOT NULL UNIQUE,
          name TEXT NOT NULL,
          unit TEXT NOT NULL,
          opening_qty REAL DEFAULT 0,
          opening_rate REAL DEFAULT 0,
          reorder_level REAL DEFAULT 0,
          is_active INTEGER DEFAULT 1,
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_purchase_bills (
          id TEXT PRIMARY KEY,
          bill_no TEXT NOT NULL,
          date_bs TEXT NOT NULL,
          supplier_name TEXT NOT NULL,
          source TEXT NOT NULL,
          source_type TEXT NOT NULL DEFAULT 'Local Purchase',
          reference_no TEXT,
          remarks TEXT,
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_purchase_lines (
          id TEXT PRIMARY KEY,
          bill_id TEXT NOT NULL,
          item_id TEXT NOT NULL,
          quantity REAL NOT NULL,
          rate REAL NOT NULL,
          amount REAL NOT NULL,
          entry_rate REAL,
          entry_amount REAL,
          FOREIGN KEY (bill_id) REFERENCES stock_purchase_bills(id),
          FOREIGN KEY (item_id) REFERENCES stock_items(id)
        );
        CREATE TABLE IF NOT EXISTS stock_sales_bills (
          id TEXT PRIMARY KEY,
          bill_no TEXT NOT NULL,
          date_bs TEXT NOT NULL,
          customer_name TEXT NOT NULL,
          source_type TEXT NOT NULL DEFAULT 'Sale',
          remarks TEXT,
          created_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS stock_sales_lines (
          id TEXT PRIMARY KEY,
          bill_id TEXT NOT NULL,
          item_id TEXT NOT NULL,
          quantity REAL NOT NULL,
          rate REAL NOT NULL,
          amount REAL NOT NULL,
          FOREIGN KEY (bill_id) REFERENCES stock_sales_bills(id),
          FOREIGN KEY (item_id) REFERENCES stock_items(id)
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_stock_items_code ON stock_items(code);
        CREATE INDEX IF NOT EXISTS idx_stock_purchase_bills_bill_no ON stock_purchase_bills(bill_no);
        CREATE INDEX IF NOT EXISTS idx_stock_purchase_bills_source_type ON stock_purchase_bills(source_type, id);
        CREATE INDEX IF NOT EXISTS idx_stock_sales_bills_bill_no ON stock_sales_bills(bill_no);
        CREATE INDEX IF NOT EXISTS idx_stock_sales_bills_source_type ON stock_sales_bills(source_type, id);
        """
    )


def load_stock(
    conn: sqlite3.Connection,
    company_id: str,
    accounts: dict[str, Any],
    purchase: dict[str, Any],
    sales_lines: list[dict[str, Any]],
    purchase_lines: list[dict[str, Any]],
    ordered_items: list[str],
) -> dict[str, Any]:
    create_stock_schema(conn)
    item_ids: dict[str, str] = {}
    for index, item in enumerate(ordered_items, start=1):
        item_id = deterministic_id("stock-item", company_id, item)
        item_ids[item] = item_id
        conn.execute(
            """
            INSERT INTO stock_items (id, code, name, unit, opening_qty, opening_rate, reorder_level, is_active, created_at)
            VALUES (?, ?, ?, 'KG', 0, 0, 0, 1, ?)
            """,
            (item_id, f"ITEM-{index:03d}", item, now_iso()),
        )

    account_parties = {party["id"]: party["name"] for party in accounts.get("parties", [])}
    sale_by_bill = {clean_bill_no(sale["billNo"]): sale for sale in accounts.get("sales", [])}
    purchase_parties = {party["id"]: party["name"] for party in purchase.get("parties", [])}
    purchase_by_bill = {clean_bill_no(row["vendorBillNumber"]): row for row in purchase.get("purchases", [])}
    sales_grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    purchase_grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for line in sales_lines:
        sales_grouped[line["billNo"]].append(line)
    for line in purchase_lines:
        purchase_grouped[line["billNo"]].append(line)

    unmatched_sales = sorted(set(sales_grouped) - set(sale_by_bill))
    unmatched_purchase = sorted(set(purchase_grouped) - set(purchase_by_bill))
    if unmatched_sales or unmatched_purchase:
        raise RuntimeError(
            "Unmatched workbook documents: "
            f"sales={unmatched_sales[:20]} purchase={unmatched_purchase[:20]}"
        )

    sales_mismatches = []
    purchase_mismatches = []
    for bill_no, lines in sales_grouped.items():
        sale = sale_by_bill[bill_no]
        bill_id = sale["id"]
        line_total = round(sum(line["amount"] for line in lines), 2)
        source_total = round(number(sale.get("salesAmount", 0)), 2)
        if abs(line_total - source_total) > 0.5:
            sales_mismatches.append((bill_no, line_total, source_total))
        conn.execute(
            """
            INSERT INTO stock_sales_bills (id, bill_no, date_bs, customer_name, source_type, remarks, created_at)
            VALUES (?, ?, ?, ?, 'Sale', ?, ?)
            """,
            (
                bill_id,
                bill_no,
                sale.get("dateBs") or lines[0]["date"],
                account_parties.get(sale.get("partyId", ""), lines[0]["partyName"]),
                sale.get("remarks", ""),
                sale.get("createdAt", now_iso()),
            ),
        )
        for index, line in enumerate(lines, start=1):
            item_id = item_ids[line["itemName"]]
            conn.execute(
                """
                INSERT INTO stock_sales_lines (id, bill_id, item_id, quantity, rate, amount)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (
                    deterministic_id("stock-sales-line", bill_id, index, line["itemName"]),
                    bill_id,
                    item_id,
                    line["quantity"],
                    line["rate"],
                    line["amount"],
                ),
            )

    for bill_no, lines in purchase_grouped.items():
        row = purchase_by_bill[bill_no]
        document_id = row["id"]
        bill_id = f"Import Purchase:{document_id}"
        landed_total = round(sum(line["amount"] for line in lines), 2)
        landed_source = round(number(row.get("landedCostNPR", 0)), 2)
        entry_total = round(sum(line["entryAmount"] for line in lines), 2)
        entry_source = round(number(row.get("amountIC", 0)), 2)
        if abs(landed_total - landed_source) > 1 or abs(entry_total - entry_source) > 1:
            purchase_mismatches.append((bill_no, landed_total, landed_source, entry_total, entry_source))
        conn.execute(
            """
            INSERT INTO stock_purchase_bills (
              id, bill_no, date_bs, supplier_name, source, source_type, reference_no, remarks, created_at
            )
            VALUES (?, ?, ?, ?, 'Importation', 'Import Purchase', ?, ?, ?)
            """,
            (
                bill_id,
                bill_no,
                row.get("debitNoteDate") or lines[0]["date"],
                purchase_parties.get(row.get("vendorPartyId", ""), lines[0]["partyName"]),
                row.get("debitNoteNumber", ""),
                row.get("remarks", ""),
                row.get("createdAt", now_iso()),
            ),
        )
        for index, line in enumerate(lines, start=1):
            item_id = item_ids[line["itemName"]]
            landed_rate = line["amount"] / line["quantity"] if line["quantity"] else 0
            conn.execute(
                """
                INSERT INTO stock_purchase_lines (
                  id, bill_id, item_id, quantity, rate, amount, entry_rate, entry_amount
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    deterministic_id("stock-purchase-line", bill_id, index, line["itemName"]),
                    bill_id,
                    item_id,
                    line["quantity"],
                    landed_rate,
                    line["amount"],
                    line["entryRate"],
                    line["entryAmount"],
                ),
            )

    conn.commit()
    return {
        "items": len(ordered_items),
        "sales_bills": len(sales_grouped),
        "sales_lines": len(sales_lines),
        "purchase_bills": len(purchase_grouped),
        "purchase_lines": len(purchase_lines),
        "sales_line_amount": round(sum(line["amount"] for line in sales_lines), 2),
        "purchase_line_landed_amount": round(sum(line["amount"] for line in purchase_lines), 2),
        "purchase_line_entry_amount": round(sum(line["entryAmount"] for line in purchase_lines), 2),
        "sales_mismatches": sales_mismatches[:20],
        "purchase_mismatches": purchase_mismatches[:20],
        "sales_mismatch_count": len(sales_mismatches),
        "purchase_mismatch_count": len(purchase_mismatches),
    }


def write_seed(profile: dict[str, Any]) -> None:
    seed_profile = {
        "companyGroupId": profile.get("companyGroupId", profile["id"]),
        "id": profile["id"],
        "isLocked": bool(profile.get("isLocked", False)),
        "lastCarryForwardAt": profile.get("lastCarryForwardAt", ""),
        "lockedAt": profile.get("lockedAt", ""),
        "name": profile["name"],
        "nextCompanyId": profile.get("nextCompanyId", ""),
        "previousCompanyId": profile.get("previousCompanyId", ""),
        "fiscalYear": profile.get("fiscalYear", ""),
        "createdAt": profile.get("createdAt", now_iso()),
        "updatedAt": profile.get("updatedAt", now_iso()),
        "trackInventory": True,
    }
    (APPDATA_DIR / "company-profiles.seed.json").write_text(
        json.dumps([seed_profile], indent=2), encoding="utf-8"
    )


def count_rows(conn: sqlite3.Connection, table: str) -> int:
    return int(conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0])


def main() -> None:
    if not BACKUP_PATH.exists():
        raise FileNotFoundError(BACKUP_PATH)
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(WORKBOOK_PATH)

    data = json.loads(BACKUP_PATH.read_text(encoding="utf-8"))
    profile = data["company"]
    company_id = profile["id"]
    fy_code = profile.get("fiscalYear") or data.get("purchase", {}).get("settings", {}).get("fiscalYear") or "2082/83"

    backup_existing_appdata()
    accounts_db = APPDATA_DIR / f"accounts-{company_id}.db"
    purchase_db = APPDATA_DIR / f"import-purchases-{company_id}.db"
    stock_db = APPDATA_DIR / stock_database_name(company_id)

    sales_lines, purchase_lines, ordered_items = parse_workbook()

    with connect(accounts_db) as conn:
        load_accounts(conn, data.get("accounts", {}), company_id, fy_code)
        account_counts = {
            "parties": count_rows(conn, "parties"),
            "sales": count_rows(conn, "sales"),
            "collections": count_rows(conn, "collections"),
            "credit_notes": count_rows(conn, "credit_notes"),
            "activity_logs": count_rows(conn, "activity_logs"),
        }

    with connect(purchase_db) as conn:
        load_purchase(conn, data.get("purchase", {}), company_id, fy_code)
        purchase_counts = {
            "parties": count_rows(conn, "parties"),
            "purchases": count_rows(conn, "import_purchases"),
            "payments": count_rows(conn, "payments"),
            "local_expenses": count_rows(conn, "local_expenses"),
            "activity_logs": count_rows(conn, "activity_logs"),
        }

    with connect(stock_db) as conn:
        stock_summary = load_stock(
            conn,
            company_id,
            data.get("accounts", {}),
            data.get("purchase", {}),
            sales_lines,
            purchase_lines,
            ordered_items,
        )
        stock_counts = {
            "items": count_rows(conn, "stock_items"),
            "purchase_bills": count_rows(conn, "stock_purchase_bills"),
            "purchase_lines": count_rows(conn, "stock_purchase_lines"),
            "sales_bills": count_rows(conn, "stock_sales_bills"),
            "sales_lines": count_rows(conn, "stock_sales_lines"),
        }

    write_seed(profile)

    print(
        json.dumps(
            {
                "appDataDir": str(APPDATA_DIR),
                "backupDir": str(BACKUP_DIR) if BACKUP_DIR.exists() else "",
                "company": profile,
                "databases": {
                    "accounts": accounts_db.name,
                    "purchase": purchase_db.name,
                    "stock": stock_db.name,
                },
                "accountCounts": account_counts,
                "purchaseCounts": purchase_counts,
                "stockCounts": stock_counts,
                "stockSummary": stock_summary,
                "seedFile": str(APPDATA_DIR / "company-profiles.seed.json"),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
